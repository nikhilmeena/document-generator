"""Generate System Test Cases spreadsheet from template and uploaded test cases."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from shutil import copyfile
from tempfile import NamedTemporaryFile

from openpyxl import load_workbook

from models.project_context import ProjectContext

TEST_CASE_SHEET = "Test Cases"
TEST_CASE_SHEET_ALIASES = (
    TEST_CASE_SHEET,
    "TestCases",
    "Test cases",
    "Sheet1",
)
OUTPUT_SHEET = "Change_Requests"
OVERWRITE_START_ROW = 10
YELLOW_FILLS = {"FFFFFF00", "FFFF00"}


def _resolve_test_case_sheet(sheetnames: list[str]) -> str:
    for name in TEST_CASE_SHEET_ALIASES:
        if name in sheetnames:
            return name

    if len(sheetnames) == 1:
        return sheetnames[0]

    available = ", ".join(sheetnames)
    raise ValueError(
        f"Uploaded test case file must contain a '{TEST_CASE_SHEET}' sheet. "
        f"Found: {available}."
    )


def _parse_date(value: str) -> datetime:
    value = value.strip()
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    raise ValueError(
        "Deployment date must be in DD-MM-YYYY format (example: 18-05-2026)."
    )


def _is_yellow_cell(cell) -> bool:
    fill = cell.fill
    if not fill or not fill.fgColor:
        return False
    rgb = (fill.fgColor.rgb or "").upper()
    return rgb in YELLOW_FILLS or rgb.endswith("FFFF00")


def _fill_yellow_cells(worksheet, context: ProjectContext) -> None:
    deployment_date = _parse_date(context.deployment_date)
    module_name = context.module_name()
    explicit_map = {
        "F4": context.cr_number,
        "C6": module_name,
        "C8": deployment_date,
        "D8": deployment_date,
    }

    for coordinate, value in explicit_map.items():
        worksheet[coordinate] = value

    for row in worksheet.iter_rows():
        for cell in row:
            if cell.coordinate in explicit_map:
                continue
            if _is_yellow_cell(cell):
                cell.value = module_name


def _clear_from_row(worksheet, start_row: int) -> None:
    max_row = worksheet.max_row or start_row
    max_col = worksheet.max_column or 1
    for row in range(start_row, max_row + 1):
        for col in range(1, max_col + 1):
            worksheet.cell(row=row, column=col).value = None


def _copy_test_case_block(source_ws, target_ws, start_row: int) -> int:
    rows_copied = 0
    max_col = source_ws.max_column or 1

    for source_row in source_ws.iter_rows(min_row=1, values_only=True):
        if not any(value is not None and str(value).strip() for value in source_row):
            continue

        target_row = start_row + rows_copied
        for col_index, value in enumerate(source_row, start=1):
            if col_index > max_col and value is None:
                continue
            target_ws.cell(row=target_row, column=col_index, value=value)
        rows_copied += 1

    return rows_copied


def generate_system_test_cases(
    template_path: Path,
    testcase_path: Path,
    context: ProjectContext,
) -> Path:
    preview_workbook = load_workbook(testcase_path, read_only=True)
    test_case_sheet = _resolve_test_case_sheet(preview_workbook.sheetnames)
    preview_workbook.close()

    with NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
        output_path = Path(temp_file.name)

    copyfile(template_path, output_path)
    workbook = load_workbook(output_path)
    if OUTPUT_SHEET not in workbook.sheetnames:
        raise ValueError(
            f"Template must contain a '{OUTPUT_SHEET}' sheet."
        )

    worksheet = workbook[OUTPUT_SHEET]
    _fill_yellow_cells(worksheet, context)
    _clear_from_row(worksheet, OVERWRITE_START_ROW)

    source_workbook = load_workbook(testcase_path, read_only=True, data_only=True)
    source_ws = source_workbook[test_case_sheet]
    rows_copied = _copy_test_case_block(source_ws, worksheet, OVERWRITE_START_ROW)
    source_workbook.close()

    if rows_copied == 0:
        raise ValueError("Uploaded test case sheet has no rows to copy.")

    workbook.save(output_path)
    return output_path
